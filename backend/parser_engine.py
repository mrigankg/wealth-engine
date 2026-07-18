import re
import hashlib
import csv
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional
import pdfplumber
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from email import message_from_file

# Built-in bank statement configuration templates
DEFAULT_TEMPLATES = {
    "HDFC Bank Savings": {
        "account_type": "BANK",
        "institution": "HDFC Bank",
        "identification_keywords": ["HDFC BANK", "STATEMENT OF ACCOUNT", "NARRATION"],
        "columns": {
            "date": 0,
            "description": 1,
            "ref_no": 2,
            "debit": 4,
            "credit": 5,
            "balance": 6
        },
        "date_format": "%d/%m/%y",
        "start_trigger": r"Date\s+Narration\s+Chq/Ref\s+No\.",
        "end_trigger": r"STATEMENT\s+TOTALS|Statement\s+Summary"
    },
    "ICICI Bank Savings": {
        "account_type": "BANK",
        "institution": "ICICI Bank",
        "identification_keywords": ["ICICI Bank", "Transaction Remarks", "Cheque No."],
        "columns": {
            "date": 1,
            "description": 3,
            "ref_no": 2,
            "debit": 4,
            "credit": 5,
            "balance": 6
        },
        "date_format": "%d/%m/%Y",
        "start_trigger": r"Value\s+Date\s+Transaction\s+Date",
        "end_trigger": r"Total\s+:"
    },
    "State Bank of India Savings": {
        "account_type": "BANK",
        "institution": "State Bank of India",
        "identification_keywords": ["STATE BANK OF INDIA", "Txn Date", "Value Date", "Description"],
        "columns": {
            "date": 0,
            "description": 2,
            "ref_no": 3,
            "debit": 4,
            "credit": 5,
            "balance": 6
        },
        "date_format": "%d\s+%b\s+%Y",  # E.g., "19 Jul 2026"
        "start_trigger": r"Txn\s+Date\s+Value\s+Date",
        "end_trigger": r"Statement\s+Summary"
    },
    "Bank of Baroda Savings": {
        "account_type": "BANK",
        "institution": "Bank of Baroda",
        "identification_keywords": ["BANK OF BARODA", "Narration", "Cheque No"],
        "columns": {
            "date": 0,
            "description": 1,
            "ref_no": 2,
            "debit": 3,
            "credit": 4,
            "balance": 5
        },
        "date_format": "%d-%m-%Y",
        "start_trigger": r"Date\s+Narration\s+Cheque\s+No",
        "end_trigger": r"Carried\s+Forward|Total"
    }
}

# ----------------- PARSING ENGINE HELPERS -----------------

def clean_amount(val: Any) -> float:
    """Cleans numeric string values from statement rows, returning a float."""
    if val is None:
        return 0.0
    val_str = str(val).strip().replace(",", "")
    if not val_str or val_str == "-" or val_str == "0":
        return 0.0
    # Strip any trailing credit/debit markers (e.g. 'Cr' or 'Dr' in SBI)
    val_str = re.sub(r"[^\d\.\-]", "", val_str)
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def parse_date(date_str: str, date_formats: List[str]) -> str:
    """Parses date string with various formats, returning standard YYYY-MM-DD format."""
    date_str = re.sub(r"\s+", " ", date_str.strip())
    # Try the user-configured list of formats
    for fmt in date_formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Default fallbacks
    for fmt in ["%d/%m/%y", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%Y-%m-%d"]:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Return original if unable to parse
    return date_str

def generate_hash(*args) -> str:
    """Generates a unique SHA-256 hash to avoid duplicate transaction imports."""
    hash_str = "-".join([str(a).strip() for a in args])
    return hashlib.sha256(hash_str.encode("utf-8")).hexdigest()

# ----------------- core parsing functions -----------------

def detect_template(first_page_text: str, custom_templates: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """Detects which bank template matches the statement based on identification keywords."""
    # Check custom user-created templates first
    for t in custom_templates:
        keywords = t.get("identification_keywords", [])
        if keywords and all(k.lower() in first_page_text.lower() for k in keywords):
            return t
            
    # Check default templates
    for name, t in DEFAULT_TEMPLATES.items():
        keywords = t.get("identification_keywords", [])
        if keywords and all(k.lower() in first_page_text.lower() for k in keywords):
            # Include name for mapping
            t_copy = t.copy()
            t_copy["name"] = name
            return t_copy
            
    return None

def parse_pdf_statement(file_path: str, template: Dict[str, Any]) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """
    Parses a PDF statement using its template configuration.
    Returns: (account_details, list_of_transactions)
    """
    cols = template["columns"]
    date_format = template["date_format"]
    start_trigger = template["start_trigger"]
    end_trigger = template["end_trigger"]
    
    transactions = []
    account_suffix = "Unknown"
    
    # 1. Extract account suffix from PDF text
    with pdfplumber.open(file_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"
        
        # Suffix matching heuristics
        account_match = re.search(r"Account\s+No\.?\s*:\s*(\d{4,})", full_text, re.IGNORECASE)
        if not account_match:
            account_match = re.search(r"A/c\s+No\.?\s*(\d{4,})", full_text, re.IGNORECASE)
        if account_match:
            account_no = account_match.group(1)
            account_suffix = account_no[-4:]
            
        # 2. Iterate pages and parse tables
        started = False
        for page_num, page in enumerate(pdf.pages):
            table = page.extract_table()
            if not table:
                continue
                
            for row in table:
                # Clean row elements (remove line breaks inside cells)
                row = [str(cell).replace("\n", " ").strip() if cell is not None else "" for cell in row]
                
                # Check for end of transaction list
                row_joined = " ".join(row)
                if started and end_trigger and re.search(end_trigger, row_joined, re.IGNORECASE):
                    break
                    
                # Check for start trigger
                if not started:
                    if re.search(start_trigger, row_joined, re.IGNORECASE):
                        started = True
                    continue
                
                # Check if row has enough columns and is a valid transaction row
                if len(row) <= max(cols.values()):
                    continue
                    
                # Extract values
                date_val = row[cols["date"]].strip()
                desc_val = row[cols["description"]].strip()
                ref_val = row[cols["ref_no"]].strip() if "ref_no" in cols and cols["ref_no"] < len(row) else ""
                
                # Validate date
                std_date = parse_date(date_val, [date_format])
                if not re.match(r"^\d{4}-\d{2}-\d{2}$", std_date):
                    continue  # Not a valid transaction line
                
                # Debit / Credit
                debit_val = clean_amount(row[cols["debit"]]) if cols["debit"] < len(row) else 0.0
                credit_val = clean_amount(row[cols["credit"]]) if cols["credit"] < len(row) else 0.0
                balance_val = clean_amount(row[cols["balance"]]) if "balance" in cols and cols["balance"] < len(row) else None
                
                if debit_val == 0.0 and credit_val == 0.0:
                    continue  # Skip row with no transaction amounts
                    
                txn_hash = generate_hash(std_date, desc_val, debit_val, credit_val, ref_val)
                
                transactions.append({
                    "date": std_date,
                    "description": desc_val,
                    "reference_no": ref_val if ref_val else None,
                    "debit": debit_val,
                    "credit": credit_val,
                    "balance": balance_val,
                    "category": "Uncategorized",
                    "transaction_hash": txn_hash
                })

    account_details = {
        "name": f"{template['institution']} Statement (*{account_suffix})",
        "institution": template["institution"],
        "account_type": template["account_type"],
        "account_number_suffix": account_suffix
    }
    
    return account_details, transactions

# ----------------- EMAIL TRANSACTION ALERTS PARSING -----------------

def parse_transaction_email(file_path: str) -> Optional[Tuple[Dict[str, Any], List[Dict[str, Any]]]]:
    """
    Parses a transaction email (.eml or .html) to extract expense transaction details.
    Uses regex pattern matchers for common banks (HDFC, ICICI, etc.) and receipt notifications.
    """
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        msg = message_from_file(f)
        
    subject = msg.get("subject", "")
    date_header = msg.get("date", "")
    
    # Extract body content (prefers HTML, fallback to text)
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            if content_type in ("text/html", "text/plain"):
                payload = part.get_payload(decode=True)
                if payload:
                    body += payload.decode("utf-8", errors="ignore")
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            body = payload.decode("utf-8", errors="ignore")
            
    # Clean HTML tags to extract raw text
    soup = BeautifulSoup(body, "html.parser")
    raw_text = soup.get_text(separator=" ")
    raw_text = re.sub(r"\s+", " ", raw_text)
    
    # Check date
    parsed_date = datetime.now().strftime("%Y-%m-%d")
    if date_header:
        # Sat, 18 Jul 2026 19:30:00 +0530
        try:
            # Strip timezone info if needed or parse standard
            date_clean = re.sub(r"\s*\(.*?\)", "", date_header)
            dt = datetime.strptime(date_clean.split(" +")[0].split(" -")[0].strip(), "%a, %d %b %Y %H:%M:%S")
            parsed_date = dt.strftime("%Y-%m-%d")
        except Exception:
            pass

    # Regex Matchers for common transaction alerts
    # 1. HDFC Card transaction alert
    hdfc_cc_match = re.search(
        r"spent Rs\.?\s*([\d\.,]+) on HDFC Bank Credit Card ending (\d+) at (.*?) on (\d{2}-\d{2}-\d{4})", 
        raw_text, re.IGNORECASE
    )
    if hdfc_cc_match:
        amount = clean_amount(hdfc_cc_match.group(1))
        card_suffix = hdfc_cc_match.group(2)
        merchant = hdfc_cc_match.group(3).strip()
        txn_date = parse_date(hdfc_cc_match.group(4), ["%d-%m-%Y"])
        
        account = {
            "name": f"HDFC Credit Card (*{card_suffix})",
            "institution": "HDFC Bank",
            "account_type": "CREDIT_CARD",
            "account_number_suffix": card_suffix
        }
        
        txn_hash = generate_hash(txn_date, merchant, amount, "HDFC_EMAIL")
        txn = {
            "date": txn_date,
            "description": f"Email Alert: Spent at {merchant}",
            "reference_no": "EMAIL_ALERT",
            "debit": amount,
            "credit": 0.0,
            "balance": None,
            "category": "Uncategorized",
            "transaction_hash": txn_hash
        }
        return account, [txn]
        
    # 2. ICICI Bank Transaction Alert
    icici_alert_match = re.search(
        r"spent INR\s*([\d\.,]+) on ICICI Bank Card ending (\d+) at (.*?) on (\d{2}-[a-zA-Z]{3}-\d{2})",
        raw_text, re.IGNORECASE
    )
    if icici_alert_match:
        amount = clean_amount(icici_alert_match.group(1))
        card_suffix = icici_alert_match.group(2)
        merchant = icici_alert_match.group(3).strip()
        # "18-Jul-26" -> %d-%b-%y
        txn_date = parse_date(icici_alert_match.group(4), ["%d-%b-%y"])
        
        account = {
            "name": f"ICICI Credit Card (*{card_suffix})",
            "institution": "ICICI Bank",
            "account_type": "CREDIT_CARD",
            "account_number_suffix": card_suffix
        }
        
        txn_hash = generate_hash(txn_date, merchant, amount, "ICICI_EMAIL")
        txn = {
            "date": txn_date,
            "description": f"Email Alert: Spent at {merchant}",
            "reference_no": "EMAIL_ALERT",
            "debit": amount,
            "credit": 0.0,
            "balance": None,
            "category": "Uncategorized",
            "transaction_hash": txn_hash
        }
        return account, [txn]
        
    # 3. Generic UPI Receipt Matcher
    upi_match = re.search(
        r"sent\s+(?:Rs\.|INR)\s*([\d\.,]+)\s+to\s+(.*?)\s+from\s+Account\s+.*(\d{4}).*UPI\s+Ref\s+(\d+)",
        raw_text, re.IGNORECASE
    )
    if upi_match:
        amount = clean_amount(upi_match.group(1))
        payee = upi_match.group(2).strip()
        bank_suffix = upi_match.group(3)
        ref_no = upi_match.group(4)
        
        account = {
            "name": f"UPI Account (*{bank_suffix})",
            "institution": "UPI Bank",
            "account_type": "BANK",
            "account_number_suffix": bank_suffix
        }
        
        txn_hash = generate_hash(parsed_date, payee, amount, ref_no)
        txn = {
            "date": parsed_date,
            "description": f"UPI Outflow: {payee}",
            "reference_no": ref_no,
            "debit": amount,
            "credit": 0.0,
            "balance": None,
            "category": "Uncategorized",
            "transaction_hash": txn_hash
        }
        return account, [txn]

    # If it is a generic Swiggy / Uber receipt html email, parse it using selectors
    if "swiggy" in subject.lower():
        # Look for swiggy order amounts
        amount_match = re.search(r"Bill Total\s+(?:Rs\.|INR|₹)\s*([\d\.,]+)", raw_text, re.IGNORECASE)
        if amount_match:
            amount = clean_amount(amount_match.group(1))
            account = {
                "name": "Cash Wallet",
                "institution": "Cash/Self",
                "account_type": "BANK",
                "account_number_suffix": "0000"
            }
            txn_hash = generate_hash(parsed_date, "Swiggy Order", amount, "SWIGGY_EMAIL")
            txn = {
                "date": parsed_date,
                "description": "Swiggy Food Order",
                "reference_no": "EMAIL_RECEIPT",
                "debit": amount,
                "credit": 0.0,
                "balance": None,
                "category": "Food & Dining",
                "transaction_hash": txn_hash
            }
            return account, [txn]

    return None

# ----------------- EXCEL & CSV GENERIC PARSERS -----------------

def parse_csv_or_excel(file_path: str, columns_mapping: Dict[str, int], date_format: str, has_headers: bool = True) -> List[Dict[str, Any]]:
    """
    Generic CSV/Excel parser that reads tabular transactions using a columns mapping dictionary.
    columns_mapping can define: date, description, ref_no, debit, credit, balance
    """
    rows = []
    
    # 1. Read files into simple matrix list
    raw_rows = []
    if file_path.endswith((".xlsx", ".xls")):
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            raw_rows.append(list(row))
    else:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                raw_rows.append(row)
                
    # 2. Parse rows
    started = False
    for i, row in enumerate(raw_rows):
        if i == 0 and has_headers:
            continue  # Skip header row
            
        if len(row) <= max(columns_mapping.values()):
            continue
            
        date_raw = str(row[columns_mapping["date"]]).strip()
        desc_raw = str(row[columns_mapping["description"]]).strip()
        ref_raw = str(row[columns_mapping["ref_no"]]).strip() if "ref_no" in columns_mapping else ""
        
        std_date = parse_date(date_raw, [date_format])
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", std_date):
            continue  # Ignore invalid date row
            
        debit = clean_amount(row[columns_mapping["debit"]]) if "debit" in columns_mapping else 0.0
        credit = clean_amount(row[columns_mapping["credit"]]) if "credit" in columns_mapping else 0.0
        balance = clean_amount(row[columns_mapping["balance"]]) if "balance" in columns_mapping else None
        
        if debit == 0.0 and credit == 0.0:
            continue
            
        txn_hash = generate_hash(std_date, desc_raw, debit, credit, ref_raw)
        
        rows.append({
            "date": std_date,
            "description": desc_raw,
            "reference_no": ref_raw if ref_raw else None,
            "debit": debit,
            "credit": credit,
            "balance": balance,
            "category": "Uncategorized",
            "transaction_hash": txn_hash
        })
        
    return rows

# ----------------- INVESTMENT STATEMENT PDF PARSER -----------------

def parse_investment_pdf_statement(file_path: str) -> Optional[Tuple[Dict[str, Any], List[Dict[str, Any]]]]:
    """
    Scans a PDF statement for Mutual Fund (CAS) or Stock ledger patterns.
    Returns: (account_details, list_of_investments_payload)
    Each investment payload contains:
      - asset_name
      - symbol_or_code
      - transactions: list of trades
    """
    full_text = ""
    tables_data = []
    
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"
            table = page.extract_table()
            if table:
                tables_data.append(table)
                
    is_mf = any(k in full_text.lower() for k in ["mutual fund", "folio no", "cams", "kfintech", "amc", "nav value"])
    is_stock = any(k in full_text.lower() for k in ["stock ledger", "contract note", "demat", "shares", "brokerage", "nse", "bse"])
    
    if not (is_mf or is_stock):
        return None
        
    # Suffix matching
    suffix = "Unknown"
    suffix_match = re.search(r"(?:Folio|Demat|Client|Account)\s*(?:No\.?\s*|\s+)?:\s*(\w{4,})", full_text, re.IGNORECASE)
    if suffix_match:
        suffix = suffix_match.group(1)[-4:]
        
    investments_map = {}  # symbol/asset_name -> {asset_name, symbol_or_code, transactions}
    
    # Simple heuristic table parsing for investment rows
    for table in tables_data:
        for row in table:
            row = [str(cell).replace("\n", " ").strip() if cell is not None else "" for cell in row]
            row_joined = " ".join(row)
            
            # Skip header or summary lines
            if "total" in row_joined.lower() or "summary" in row_joined.lower():
                continue
                
            # We look for rows that contain:
            # 1. A date
            # 2. A transaction type (BUY, SELL, PURCHASE, REDEMPTION, etc.)
            # 3. Numeric values for quantity, price/NAV, and amount
            
            # Find date in row
            date_val = None
            date_idx = -1
            for idx, cell in enumerate(row):
                std_date = parse_date(cell, [])
                if re.match(r"^\d{4}-\d{2}-\d{2}$", std_date):
                    date_val = std_date
                    date_idx = idx
                    break
                    
            if not date_val:
                continue
                
            # Find transaction type
            tx_type = "BUY"
            if any(k in row_joined.lower() for k in ["sell", "redemption", "payout", "redempt"]):
                tx_type = "SELL"
            elif "dividend" in row_joined.lower() or "reinvest" in row_joined.lower():
                tx_type = "DIVIDEND_REINVEST"
                
            # Try to identify asset name from text preceding or within row
            # For simplicity in statements, search for columns that contain non-numeric words (asset name)
            asset_name = "Unknown Asset"
            for cell in row:
                if len(cell) > 8 and not re.search(r"\d", cell) and any(k in cell.lower() for k in ["fund", "growth", "equity", "limited", "ltd", "co"]):
                    asset_name = cell
                    break
            
            if asset_name == "Unknown Asset":
                # Fallback: check if we can extract it from the text near the table or row
                # In CAS statement, fund names are often on a separate line above tables.
                # Let's extract any words that look like a fund/stock name.
                fund_names = re.findall(r"([A-Z][A-Za-z0-9\s\-]+(?:Fund|Growth|Equity|Limited|Ltd|Co|Direct))", full_text)
                if fund_names:
                    asset_name = fund_names[0].strip()
                    
            symbol = asset_name.replace(" ", "_").upper()
            
            # Look for quantity, price, total amount
            # Usually: Quantity (units/shares), Price (NAV/share price), Total Amount
            # We filter numbers in the row excluding the date index
            numbers = []
            for idx, cell in enumerate(row):
                if idx == date_idx:
                    continue
                num = clean_amount(cell)
                if num > 0:
                    numbers.append((idx, num))
                    
            # Need at least 2 numbers (e.g. quantity and price, or quantity and amount)
            if len(numbers) < 2:
                continue
                
            # Sort numbers by value. Usually, Price is moderate, Quantity is variable, Amount is largest.
            # E.g., numbers: [(idx1, 50.0), (idx2, 10.0), (idx3, 500.0)]
            # Let's do a logical mapping:
            # Let's check if any combination of numA * numB approximates numC (within 1% margin)
            qty = 0.0
            price = 0.0
            amt = 0.0
            
            sorted_nums = sorted(numbers, key=lambda x: x[1])
            if len(sorted_nums) >= 3:
                # E.g. [price, qty, amount] or [qty, price, amount]
                n1, n2, n3 = sorted_nums[0][1], sorted_nums[1][1], sorted_nums[2][1]
                if abs(n1 * n2 - n3) < (0.02 * n3):
                    qty = n2
                    price = n1
                    amt = n3
                elif abs(n1 * n3 - n2) < (0.02 * n2): # shouldn't happen usually
                    qty = n1
                    price = n3
                    amt = n2
            elif len(sorted_nums) == 2:
                # E.g. Quantity and Price, calculate Amount
                n1, n2 = sorted_nums[0][1], sorted_nums[1][1]
                # If we have NAV/Price, it's usually smaller than total amount.
                # Let's assume n1 is price/qty and n2 is amount, or calculate:
                # Let's just assume price is smaller than amount, quantity is amount/price.
                # Since we only have 2 numbers, let's treat n1 as quantity, n2 as price, amt = n1 * n2.
                qty = n1
                price = n2
                amt = qty * price
                
            if qty == 0.0 or price == 0.0:
                continue
                
            txn_hash = generate_hash(date_val, asset_name, qty, price, tx_type)
            
            trade = {
                "date": date_val,
                "transaction_type": tx_type,
                "quantity": qty,
                "price_per_unit": price,
                "total_amount": amt,
                "transaction_hash": txn_hash
            }
            
            if symbol not in investments_map:
                investments_map[symbol] = {
                    "asset_name": asset_name,
                    "symbol_or_code": symbol,
                    "transactions": []
                }
            investments_map[symbol]["transactions"].append(trade)
            
    if not investments_map:
        return None
        
    account_details = {
        "name": f"Portfolio Account (*{suffix})",
        "institution": "CAMS CAS" if is_mf else "Zerodha Ledger",
        "account_type": "MUTUAL_FUND" if is_mf else "STOCK",
        "account_number_suffix": suffix
    }
    
    return account_details, list(investments_map.values())

