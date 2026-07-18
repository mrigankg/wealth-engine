import os
import shutil
from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from typing import List, Optional
import json

from backend import db, parser_engine, models

app = FastAPI(title="Offline Personal Finance Tracker", version="1.0.0")

# Enable CORS for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize Database on app startup
@app.on_event("startup")
def startup_event():
    db.init_db()

# ----------------- ACCOUNTS & SUMMARY -----------------

@app.get("/api/accounts")
def get_accounts():
    try:
        return db.get_accounts()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/summary")
def get_financial_summary():
    """Calculates balances, total holdings, debt, and returns a high-level summary."""
    conn = db.get_db_connection()
    try:
        cursor = conn.cursor()
        
        # 1. Bank Account balances (SUM of credit - debit per account)
        # Note: If statement balance is provided in latest transaction, we could use it, 
        # but summing deposits - withdrawals is the standard ledger approach.
        cursor.execute(
            """
            SELECT 
                a.id, a.name, a.institution, a.account_type, a.account_number_suffix,
                COALESCE(SUM(t.credit), 0) - COALESCE(SUM(t.debit), 0) as calculated_balance
            FROM accounts a
            LEFT JOIN bank_transactions t ON a.id = t.account_id
            WHERE a.account_type IN ('BANK', 'CREDIT_CARD')
            GROUP BY a.id
            """
        )
        bank_accounts = [dict(row) for row in cursor.fetchall()]
        
        total_cash = 0.0
        total_cc_debt = 0.0
        
        for acc in bank_accounts:
            bal = acc["calculated_balance"]
            if acc["account_type"] == "BANK":
                total_cash += bal
            elif acc["account_type"] == "CREDIT_CARD":
                # Credit card balance is usually negative (spent money)
                # We show it as debt (positive value of how much is owed)
                total_cc_debt += max(0.0, -bal)

        # 2. Fixed Deposits
        cursor.execute(
            """
            SELECT COALESCE(SUM(principal_amount), 0) as total_fds 
            FROM fixed_deposits 
            WHERE status = 'ACTIVE'
            """
        )
        total_fds = cursor.fetchone()["total_fds"] or 0.0

        # 3. Investments (Mutual Funds and Stocks)
        cursor.execute(
            """
            SELECT 
                COALESCE(SUM(total_quantity * average_price), 0) as total_cost,
                COALESCE(SUM(total_quantity * CASE WHEN current_price > 0 THEN current_price ELSE average_price END), 0) as total_value
            FROM investment_holdings
            """
        )
        inv_row = cursor.fetchone()
        total_inv_cost = inv_row["total_cost"] or 0.0
        total_inv_value = inv_row["total_value"] or 0.0
        inv_gain_loss = total_inv_value - total_inv_cost

        # 4. Gold Holdings value
        cursor.execute(
            """
            SELECT COALESCE(SUM(weight_grams * current_price_per_gram), 0) as total_gold
            FROM gold_holdings
            """
        )
        total_gold = cursor.fetchone()["total_gold"] or 0.0

        # 5. Provident Funds balance
        cursor.execute(
            """
            SELECT COALESCE(SUM(current_balance), 0) as total_pf
            FROM provident_funds
            """
        )
        total_pf = cursor.fetchone()["total_pf"] or 0.0

        # 6. Real Estate equity (valuation - associated loans)
        cursor.execute(
            """
            SELECT 
                COALESCE(SUM(current_estimated_value), 0) as total_re_val,
                COALESCE(SUM(associated_loan_amount), 0) as total_re_loans
            FROM real_estate
            """
        )
        re_row = cursor.fetchone()
        total_re_val = re_row["total_re_val"] or 0.0
        total_re_loans = re_row["total_re_loans"] or 0.0
        total_re_equity = total_re_val - total_re_loans

        # 7. Insurance Policy Sum Assured
        cursor.execute(
            """
            SELECT COALESCE(SUM(sum_assured), 0) as total_insurance_cover
            FROM insurance_policies
            WHERE status = 'ACTIVE'
            """
        )
        total_insurance_cover = cursor.fetchone()["total_insurance_cover"] or 0.0

        net_worth = total_cash + total_fds + total_inv_value + total_gold + total_pf + total_re_equity - total_cc_debt

        return {
            "net_worth": net_worth,
            "total_cash": total_cash,
            "total_cc_debt": total_cc_debt,
            "total_fixed_deposits": total_fds,
            "total_investments_cost": total_inv_cost,
            "total_investments_value": total_inv_value,
            "investments_gain_loss": inv_gain_loss,
            "total_gold_value": total_gold,
            "total_pf_value": total_pf,
            "total_real_estate_value": total_re_val,
            "total_real_estate_equity": total_re_equity,
            "total_real_estate_loans": total_re_loans,
            "total_insurance_cover": total_insurance_cover,
            "bank_accounts": bank_accounts
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# ----------------- TRANSACTION QUERIES -----------------

@app.get("/api/transactions")
def get_transactions(
    account_id: Optional[int] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 200
):
    """Retrieves ledger transactions for bank and credit card accounts."""
    conn = db.get_db_connection()
    try:
        query = """
            SELECT t.*, a.name as account_name, a.institution, a.account_type
            FROM bank_transactions t
            JOIN accounts a ON t.account_id = a.id
            WHERE 1=1
        """
        params = []
        
        if account_id is not None:
            query += " AND t.account_id = ?"
            params.append(account_id)
            
        if category:
            query += " AND t.category = ?"
            params.append(category)
            
        if search:
            query += " AND (t.description LIKE ? OR t.reference_no LIKE ?)"
            params.extend([f"%{search}%", f"%{search}%"])
            
        query += " ORDER BY t.date DESC, t.id DESC LIMIT ?"
        params.append(limit)
        
        cursor = conn.cursor()
        cursor.execute(query, params)
        return [dict(row) for row in cursor.fetchall()]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.get("/api/holdings")
def get_holdings(account_id: Optional[int] = None):
    """Retrieves consolidated investments portfolio along with individual trades (child details)."""
    conn = db.get_db_connection()
    try:
        query = """
            SELECT h.*, a.name as account_name, a.institution
            FROM investment_holdings h
            JOIN accounts a ON h.account_id = a.id
            WHERE 1=1
        """
        params = []
        if account_id is not None:
            query += " AND h.account_id = ?"
            params.append(account_id)
            
        cursor = conn.cursor()
        cursor.execute(query, params)
        holdings = [dict(row) for row in cursor.fetchall()]
        
        # Load trades for each holding
        for h in holdings:
            cursor.execute(
                """
                SELECT * FROM investment_transactions 
                WHERE holding_id = ? 
                ORDER BY date DESC
                """,
                (h["id"],)
            )
            h["trades"] = [dict(row) for row in cursor.fetchall()]
            
        return holdings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.get("/api/fixed-deposits")
def get_fixed_deposits(account_id: Optional[int] = None):
    conn = db.get_db_connection()
    try:
        query = """
            SELECT fd.*, a.name as account_name, a.institution
            FROM fixed_deposits fd
            JOIN accounts a ON fd.account_id = a.id
            WHERE 1=1
        """
        params = []
        if account_id is not None:
            query += " AND fd.account_id = ?"
            params.append(account_id)
            
        cursor = conn.cursor()
        cursor.execute(query, params)
        return [dict(row) for row in cursor.fetchall()]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# ----------------- CONFIGURABLE TEMPLATES -----------------

@app.get("/api/templates")
def get_templates():
    try:
        return db.get_templates()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/templates")
def create_template(tmpl: models.TemplateSchema):
    try:
        db.save_template(tmpl.name, tmpl.config)
        return {"status": "success", "message": f"Template '{tmpl.name}' saved successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ----------------- UPLOAD & PARSING -----------------

@app.post("/api/upload")
async def upload_statement(file: UploadFile = File(...)):
    """
    Accepts statement uploads, runs identification heuristics to select a parser template,
    extracts the transactions/investments/FD details, and returns a preview structure.
    No database writes occur during this step.
    """
    # 1. Create a temporary folder and save the uploaded file locally
    temp_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "temp_uploads")
    os.makedirs(temp_dir, exist_ok=True)
    
    file_path = os.path.join(temp_dir, file.filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    try:
        file_ext = os.path.splitext(file.filename)[1].lower()
        
        # A. Handle EML / HTML Emails
        if file_ext in (".eml", ".html", ".htm"):
            result = parser_engine.parse_transaction_email(file_path)
            if not result:
                raise HTTPException(
                    status_code=400, 
                    detail="Could not identify expense or transaction alerts in this email. Make sure it contains transaction text."
                )
            account, transactions = result
            return {
                "file_type": "email_alert",
                "account": account,
                "bank_transactions": transactions
            }
            
        # B. Handle PDFs (Bank statements or Investment portfolios)
        elif file_ext == ".pdf":
            # Extract first page text for template matching
            first_page_text = ""
            with pdfplumber.open(file_path) as pdf:
                if len(pdf.pages) > 0:
                    first_page_text = pdf.pages[0].extract_text() or ""
            
            # Fetch custom templates from DB
            custom_templates = [t["config"] for t in db.get_templates()]
            
            # Try to match template
            template = parser_engine.detect_template(first_page_text, custom_templates)
            if template:
                # Bank Statement
                account, transactions = parser_engine.parse_pdf_statement(file_path, template)
                return {
                    "file_type": "bank_statement",
                    "account": account,
                    "bank_transactions": transactions
                }
            
            # If not a bank statement, try parsing as Investment (Mutual Fund CAS or Stocks)
            inv_result = parser_engine.parse_investment_pdf_statement(file_path)
            if inv_result:
                account, investments = inv_result
                return {
                    "file_type": "investment_statement",
                    "account": account,
                    "investments": investments
                }
                
            raise HTTPException(
                status_code=400,
                detail="PDF statement type not recognized. Please configure a custom template or upload a supported statement."
            )
            
        # C. Handle CSV/Excel (Needs manual column selection or template)
        elif file_ext in (".csv", ".xlsx", ".xls"):
            # For simplicity, we search if the user has saved a custom template with the filename/sheet keyword
            # E.g., if there's a custom template named matching this upload. Otherwise, we default to HDFC columns.
            # We will return the first 10 rows to let the user select/map columns, or use a default mapping
            default_mapping = {
                "date": 0,
                "description": 1,
                "ref_no": 2,
                "debit": 3,
                "credit": 4,
                "balance": 5
            }
            
            # Read sample rows
            sample_rows = []
            if file_ext == ".csv":
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    reader = csv.reader(f)
                    for idx, r in enumerate(reader):
                        if idx < 10:
                            sample_rows.append(r)
            else:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, read_only=True)
                sheet = wb.active
                for idx, r in enumerate(sheet.iter_rows(values_only=True)):
                    if idx < 10:
                        sample_rows.append([str(c) if c is not None else "" for c in r])
                        
            # Return sample rows so frontend can dynamically configure or proceed with default
            # For quick start, we parse with default mapping
            txns = parser_engine.parse_csv_or_excel(file_path, default_mapping, "%Y-%m-%d")
            
            account = {
                "name": f"Tabular Import ({file.filename})",
                "institution": "CSV/Excel Upload",
                "account_type": "BANK",
                "account_number_suffix": "CSV"
            }
            
            return {
                "file_type": "csv_excel",
                "account": account,
                "bank_transactions": txns,
                "sample_rows": sample_rows
            }
            
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please upload PDF, CSV, Excel, or EML files.")
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Parsing failed: {str(e)}")
    finally:
        # Cleanup temporary file
        if os.path.exists(file_path):
            os.remove(file_path)

# ----------------- IMPORT CONFIRMATION -----------------

@app.post("/api/confirm-import")
def confirm_import(req: models.ImportConfirmationRequest):
    """
    Saves the verified preview transactions into the SQLite database.
    Creates the account if it doesn't already exist.
    """
    try:
        # 1. Get or create Account
        account_id = db.create_account(
            name=req.account.name,
            institution=req.account.institution,
            account_type=req.account.account_type,
            account_number_suffix=req.account.account_number_suffix
        )
        
        if not account_id:
            raise Exception("Failed to resolve or create account.")
            
        imported_count = 0
        
        # 2. Insert transactions based on account type
        # A. Bank Transactions
        if req.bank_transactions and req.account.account_type in ('BANK', 'CREDIT_CARD'):
            # Convert schema objects to tuples for batch insertion
            txns_tuples = []
            for t in req.bank_transactions:
                # Generate transaction hash if not present
                txn_hash = t.transaction_hash or parser_engine.generate_hash(
                    t.date, t.description, t.debit, t.credit, t.reference_no
                )
                txns_tuples.append((
                    account_id, t.date, t.description, t.reference_no,
                    t.debit, t.credit, t.balance, t.category or "Uncategorized", txn_hash
                ))
            imported_count = db.insert_bank_transactions(txns_tuples)
            
        # B. Fixed Deposits
        elif req.fixed_deposits and req.account.account_type == 'FIXED_DEPOSIT':
            fds_tuples = []
            for fd in req.fixed_deposits:
                fds_tuples.append((
                    account_id, fd.fd_number, fd.principal_amount, fd.maturity_amount,
                    fd.interest_rate, fd.deposit_date, fd.maturity_date, fd.status
                ))
            imported_count = db.insert_fixed_deposits(fds_tuples)
            
        # C. Investments (Mutual Funds / Stocks)
        elif req.investments and req.account.account_type in ('MUTUAL_FUND', 'STOCK'):
            # Restructure investments schema for DB insertion
            txns_list = []
            for inv in req.investments:
                for t in inv.transactions:
                    txn_hash = t.transaction_hash or parser_engine.generate_hash(
                        t.date, inv.asset_name, t.quantity, t.price_per_unit, t.transaction_type
                    )
                    txns_list.append({
                        "account_id": account_id,
                        "asset_name": inv.asset_name,
                        "symbol_or_code": inv.symbol_or_code,
                        "date": t.date,
                        "transaction_type": t.transaction_type,
                        "quantity": t.quantity,
                        "price_per_unit": t.price_per_unit,
                        "total_amount": t.total_amount,
                        "transaction_hash": txn_hash
                    })
            imported_count = db.insert_investment_transactions(txns_list)
            
        return {
            "status": "success",
            "imported_count": imported_count,
            "message": f"Successfully imported {imported_count} new entries."
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

# ----------------- ACCOUNTS WRITE -----------------

@app.post("/api/accounts")
def create_account(acc: models.AccountBase):
    try:
        acc_id = db.create_account(
            name=acc.name,
            institution=acc.institution,
            account_type=acc.account_type,
            account_number_suffix=acc.account_number_suffix
        )
        return {"status": "success", "account_id": acc_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ----------------- NEW EXPANDED ASSET CLASSES ENDPOINTS -----------------

# Gold
@app.get("/api/gold")
def get_gold():
    try:
        return db.get_gold_holdings()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/gold")
def add_gold(req: models.GoldHoldingSchema):
    try:
        db.save_gold_holding(
            account_id=req.account_id,
            gold_type=req.gold_type,
            weight_grams=req.weight_grams,
            purity_carats=req.purity_carats,
            invested_amount=req.invested_amount,
            current_price_per_gram=req.current_price_per_gram
        )
        return {"status": "success", "message": "Gold holding saved."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Provident Funds
@app.get("/api/provident-funds")
def get_provident_funds():
    try:
        return db.get_provident_funds()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/provident-funds")
def add_provident_fund(req: models.ProvidentFundSchema):
    try:
        db.save_provident_fund(
            account_id=req.account_id,
            pf_type=req.pf_type,
            current_balance=req.current_balance,
            monthly_contribution=req.monthly_contribution,
            interest_rate=req.interest_rate,
            last_updated_date=req.last_updated_date
        )
        return {"status": "success", "message": "Provident fund details saved."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Real Estate
@app.get("/api/real-estate")
def get_real_estate():
    try:
        return db.get_real_estate()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/real-estate")
def add_real_estate(req: models.RealEstateSchema):
    try:
        db.save_real_estate(
            account_id=req.account_id,
            property_name=req.property_name,
            purchase_price=req.purchase_price,
            current_estimated_value=req.current_estimated_value,
            monthly_rental_income=req.monthly_rental_income,
            associated_loan_amount=req.associated_loan_amount,
            status=req.status
        )
        return {"status": "success", "message": "Real Estate property saved."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Insurance
@app.get("/api/insurance")
def get_insurance():
    try:
        return db.get_insurance_policies()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/insurance")
def add_insurance(req: models.InsurancePolicySchema):
    try:
        db.save_insurance_policy(
            account_id=req.account_id,
            policy_number=req.policy_number,
            policy_name=req.policy_name,
            policy_type=req.policy_type,
            sum_assured=req.sum_assured,
            premium_amount=req.premium_amount,
            premium_frequency=req.premium_frequency,
            due_date=req.due_date,
            maturity_date=req.maturity_date,
            status=req.status
        )
        return {"status": "success", "message": "Insurance policy details saved."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ----------------- MANUAL INPUTS FOR EXISTING ASSETS -----------------

@app.post("/api/transactions")
def add_manual_transaction(req: models.ManualTransactionRequest):
    try:
        txn_hash = parser_engine.generate_hash(req.date, req.description, req.debit, req.credit, req.reference_no, "MANUAL")
        txn = (req.account_id, req.date, req.description, req.reference_no, req.debit, req.credit, None, req.category, txn_hash)
        count = db.insert_bank_transactions([txn])
        if count == 0:
            raise Exception("Transaction already exists in database (duplicate hash).")
        return {"status": "success", "message": "Manual transaction added."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/fixed-deposits")
def add_manual_fd(req: models.ManualFixedDepositRequest):
    try:
        fd = (req.account_id, req.fd_number, req.principal_amount, req.maturity_amount, req.interest_rate, req.deposit_date, req.maturity_date, req.status)
        count = db.insert_fixed_deposits([fd])
        if count == 0:
            raise Exception("Failed to insert Fixed Deposit. Check for duplicate FD number.")
        return {"status": "success", "message": "Fixed Deposit added."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/investments/trade")
def add_manual_trade(req: models.ManualInvestmentTradeRequest):
    try:
        txn_hash = parser_engine.generate_hash(req.date, req.asset_name, req.quantity, req.price_per_unit, req.transaction_type, "MANUAL")
        trade = {
            "account_id": req.account_id,
            "asset_name": req.asset_name,
            "symbol_or_code": req.symbol_or_code,
            "date": req.date,
            "transaction_type": req.transaction_type,
            "quantity": req.quantity,
            "price_per_unit": req.price_per_unit,
            "total_amount": req.total_amount,
            "transaction_hash": txn_hash
        }
        count = db.insert_investment_transactions([trade])
        if count == 0:
            raise Exception("Trade already exists in database (duplicate hash).")
        return {"status": "success", "message": "Manual investment trade added."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ----------------- SERVE FRONTEND (OFFLINE) -----------------

# Path to the frontend directory
FRONTEND_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "frontend")

@app.get("/")
def read_root():
    """Serves the index.html on root access."""
    index_path = os.path.join(FRONTEND_DIR, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return {"message": "Server is running. Frontend directory is empty or not created yet."}

# Mount the static files (styles, scripts, local assets)
if os.path.exists(FRONTEND_DIR):
    app.mount("/", StaticFiles(directory=FRONTEND_DIR), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("backend.main:app", host="127.0.0.1", port=8000, reload=True)
